import pandas as pd
import numpy as np
import joblib
import requests
from datetime import datetime
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Import de l'utilitaire PyInstaller pour gérer les chemins
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from pyinstaller_utils import get_base_path, get_resource_path, is_pyinstaller
except ImportError:
    def get_base_path():
        return Path(__file__).parent.parent
    def get_resource_path(relative_path):
        base = Path(__file__).parent.parent
        # Chercher dans data/ ou models/ selon le type de fichier
        if relative_path.endswith('.xlsx') or relative_path.endswith('.csv'):
            data_path = base / "data" / relative_path
            if data_path.exists():
                return data_path
        elif relative_path.endswith('.pkl'):
            model_path = base / "models" / relative_path
            if model_path.exists():
                return model_path
        # Fallback: chercher à la racine
        return base / relative_path
    def is_pyinstaller():
        return False

# === PARAMÈTRES GÉNÉRAUX ===
BASE_PATH = get_base_path()
# Utiliser config.py si disponible
try:
    from config import EXCEL_PATH, MODEL_PATH, WEATHER_PATH as METEO_PATH
except ImportError:
    EXCEL_PATH = str(get_resource_path("recoltes_fraises.xlsx"))
    MODEL_PATH = str(get_resource_path("model_fraises_v2.pkl"))
    METEO_PATH = str(get_resource_path("meteo_dataset.csv"))
FORECASTS_DIR = BASE_PATH / "forecasts"  # Dans le dossier de l'exécutable, pas dans _internal
LAT, LON = 43.12, 6.14  # Hyères
FORECAST_DAYS = 7
TIMEZONE = "Europe/Paris"

# === INIT ===
print("🌤️ Génération des prévisions de récolte (1 semaine)...")
FORECASTS_DIR.mkdir(exist_ok=True)

# === CHARGEMENT DU MODÈLE ===
if not Path(MODEL_PATH).exists():
    raise FileNotFoundError(f"❌ Modèle introuvable : {MODEL_PATH}")
model = joblib.load(MODEL_PATH)
print(f"✅ Modèle chargé ({MODEL_PATH})")

# === CHARGEMENT DES PARAMÈTRES ===
# Utiliser data_loader pour compatibilité SQLite/Excel
try:
    from data_loader import load_parametres, load_jour_courant, load_recoltes, load_recolte_quotidienne, load_plants_par_annee
    USE_DATA_LOADER = True
except ImportError:
    USE_DATA_LOADER = False
    print("⚠️ Module data_loader non disponible, utilisation d'Excel uniquement")

if USE_DATA_LOADER:
    params = load_parametres()
else:
    params = pd.read_excel(EXCEL_PATH, sheet_name="Paramètres")
if params.empty:
    raise ValueError("❌ Onglet 'Paramètres' vide ou introuvable dans recoltes_fraises.xlsx.")
print(f"📘 {len(params)} combinaisons parcelle/variété chargées.")

# Vérifie si des colonnes saison sont présentes
has_season_columns = {"saison_debut", "saison_fin"}.issubset(params.columns)

# === DONNÉES JOUR COURANT ===
# Note: Le jour courant peut être vide le matin, c'est normal
if USE_DATA_LOADER:
    df_jour = load_jour_courant()
else:
    df_jour = pd.read_excel(EXCEL_PATH, sheet_name="Jour_courant", parse_dates=["date"])
if df_jour.empty:
    print("ℹ️ Aucun enregistrement dans 'Jour_courant' (normal le matin avant les récoltes).")
    print("   Les prévisions utiliseront uniquement les données historiques saisonnières.")
    df_jour = pd.DataFrame(columns=["date", "variety", "kg_premiere_rangee"])

# Fusionne avec paramètres sur variety uniquement (parcelle déduite depuis Paramètres)
if not df_jour.empty and "variety" in df_jour.columns:
    df_jour["variety"] = df_jour["variety"].astype(str).str.strip().str.lower()
    df_jour = df_jour.merge(params, on=["variety"], how="left")
    df_jour["kg_par_rangee"] = df_jour["kg_premiere_rangee"].fillna(0)
    print(f"✅ Données du jour courant chargées : {len(df_jour)} variétés")
else:
    df_jour = pd.DataFrame(columns=["date", "parcelle", "variety", "kg_premiere_rangee", "kg_par_rangee"])

# === LECTURE DE L'HISTORIQUE ===
if USE_DATA_LOADER:
    df_hist = load_recoltes()
else:
    df_hist = pd.read_excel(EXCEL_PATH, sheet_name="Recoltes", parse_dates=["date"])
# Ajouter parcelle depuis Paramètres si pas déjà présente
if "parcelle" not in df_hist.columns:
    df_hist["variety"] = df_hist["variety"].astype(str).str.strip().str.lower()
    df_hist = df_hist.merge(params[["variety", "parcelle"]], on=["variety"], how="left")
df_hist["year"] = df_hist["date"].dt.year
df_hist["month"] = df_hist["date"].dt.month

today = pd.Timestamp.now()
current_year = today.year
current_month = today.month

# === DÉTECTION DES SAISONS ===
if has_season_columns:
    print("📆 Utilisation des colonnes 'saison_debut' / 'saison_fin' depuis Paramètres.")
else:
    print("📊 Détection automatique de la saison active à partir des historiques...")
    activity = df_hist.groupby("month")["kg_total"].mean()
    SEASON_MONTHS = activity[activity > activity.mean() * 0.3].index.tolist()
    print(f"🌱 Mois actifs détectés : {SEASON_MONTHS}")

# === DÉTERMINATION DU CONTEXTE (nouvelle saison ou continuité) ===
# Ajouter nb_rangees depuis Paramètres si pas déjà présent
if "nb_rangees" not in df_hist.columns:
    df_hist = df_hist.merge(params[["variety", "nb_rangees"]], on=["variety"], how="left")
    # Gérer le cas où nb_rangees est vide
    df_hist["nb_rangees"] = df_hist["nb_rangees"].fillna(10)  # Valeur par défaut

last_values = (
    df_hist.sort_values("date")
    .groupby(["parcelle", "variety"])
    .last()
    .reset_index()[["parcelle", "variety", "date", "kg_total", "nb_rangees"]]
)
last_values["year"] = last_values["date"].dt.year
last_values["month"] = last_values["date"].dt.month
# Calcul de kg_par_rangee_prev_day (gère le cas où nb_rangees pourrait être 0)
last_values["kg_par_rangee_prev_day"] = last_values["kg_total"] / last_values["nb_rangees"].replace(0, np.nan)
last_values = last_values.dropna(subset=["kg_par_rangee_prev_day"])

def is_same_season(row):
    if has_season_columns:
        # Recherche sur variety uniquement (parcelle déduite depuis Paramètres)
        row_param = params.loc[params["variety"] == row["variety"]].iloc[0]
        # Vérifier si les valeurs de saison sont présentes et valides
        start_val = row_param["saison_debut"]
        end_val = row_param["saison_fin"]
        
        # Si les valeurs sont manquantes (None ou NaN), ne pas utiliser cette ligne pour le contexte
        if pd.isna(start_val) or pd.isna(end_val) or start_val is None or end_val is None:
            return False
        
        try:
            start, end = int(start_val), int(end_val)
            in_current = start <= current_month <= end
            in_last = start <= row["month"] <= end
            return row["year"] == current_year and in_current and in_last
        except (ValueError, TypeError):
            # Si la conversion échoue, ne pas utiliser cette ligne pour le contexte
            return False
    else:
        return (row["year"] == current_year) and (row["month"] in SEASON_MONTHS) and (current_month in SEASON_MONTHS)

last_values["keep_for_context"] = last_values.apply(is_same_season, axis=1)
valid_last = last_values[last_values["keep_for_context"]]

if valid_last.empty:
    print("🌱 Nouvelle saison détectée : contexte remis à zéro.")
else:
    print(f"📈 Continuité saison détectée : {len(valid_last)} récoltes récentes utilisées.")

# === MÉTÉO (Open-Meteo API) ===
print("📡 Téléchargement des prévisions météo...")
url = (
    f"https://api.open-meteo.com/v1/forecast?"
    f"latitude={LAT}&longitude={LON}"
    "&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,"
    "sunshine_duration,relative_humidity_2m_mean,shortwave_radiation_sum&timezone=" + TIMEZONE
)

df_forecast = None

try:
    response = requests.get(url, timeout=15)
    response.raise_for_status()
    data = response.json()
    if "daily" not in data:
        raise ValueError("Réponse météo invalide (pas de clé 'daily').")
    # Gérer shortwave_radiation_sum qui peut ne pas être disponible dans toutes les régions
    shortwave_data = data["daily"].get("shortwave_radiation_sum")
    if shortwave_data is None:
        shortwave_data = [np.nan] * len(data["daily"]["time"])
    
    df_forecast = pd.DataFrame({
        "date": pd.to_datetime(data["daily"]["time"]),
        "temp_max": data["daily"]["temperature_2m_max"],
        "temp_min": data["daily"]["temperature_2m_min"],
        "rain_mm": data["daily"]["precipitation_sum"],
        "sun_hours": np.array(data["daily"]["sunshine_duration"]) / 3600.0,
        "humidity": data["daily"]["relative_humidity_2m_mean"],
        "shortwave_radiation": shortwave_data  # W/m² (rayonnement solaire global)
    })
    print("✅ Données météo récupérées depuis l'API Open-Meteo.")
except Exception as e:
    print(f"⚠️ Impossible de récupérer les données météo en ligne : {e}")
    print("   Utilisation d'une estimation locale depuis meteo_dataset.csv ...")
    try:
        df_meteo = pd.read_csv(METEO_PATH)
        if df_meteo.empty:
            raise ValueError("meteo_dataset.csv est vide")
        # Nettoyer et prendre la moyenne des 30 derniers jours (ou tous si moins)
        df_recent = df_meteo.tail(30)
        for col in ["temp_max", "temp_min", "rain_mm", "sun_hours", "humidity", "shortwave_radiation"]:
            if col not in df_recent.columns:
                df_recent[col] = np.nan
        avg_values = df_recent.mean(numeric_only=True)
        fallback_dates = pd.date_range(start=datetime.now().date(), periods=FORECAST_DAYS, freq="D")
        df_forecast = pd.DataFrame({
            "date": fallback_dates,
            "temp_max": [avg_values.get("temp_max")] * FORECAST_DAYS,
            "temp_min": [avg_values.get("temp_min")] * FORECAST_DAYS,
            "rain_mm": [avg_values.get("rain_mm", 0)] * FORECAST_DAYS,
            "sun_hours": [avg_values.get("sun_hours", 0)] * FORECAST_DAYS,
            "humidity": [avg_values.get("humidity")] * FORECAST_DAYS,
            "shortwave_radiation": [avg_values.get("shortwave_radiation", np.nan)] * FORECAST_DAYS
        })
        print("✅ Données météo générées à partir des moyennes locales.")
    except Exception as fallback_error:
        raise RuntimeError(
            f"Impossible d'obtenir des données météo en ligne et localement : {fallback_error}"
        )
df_forecast["temp_mean"] = (df_forecast["temp_max"] + df_forecast["temp_min"]) / 2
df_forecast = df_forecast[df_forecast["date"] >= pd.Timestamp(datetime.now().date())].head(FORECAST_DAYS)

# === FILTRAGE DES VARIÉTÉS EN SAISON DE PLANTATION ===
def is_variety_in_season(row):
    """Vérifie si une variété est en saison de plantation pour le mois actuel."""
    if has_season_columns:
        start_val = row.get("saison_debut")
        end_val = row.get("saison_fin")
        
        # Si les valeurs sont manquantes, on considère que la variété n'est pas en saison
        if pd.isna(start_val) or pd.isna(end_val) or start_val is None or end_val is None:
            return False
        
        try:
            start, end = int(start_val), int(end_val)
            # Gérer les saisons qui chevauchent l'année (ex: 11-3 pour nov-mars)
            if start <= end:
                # Saison normale (ex: 3-9 pour mars-septembre)
                return start <= current_month <= end
            else:
                # Saison qui chevauche l'année (ex: 11-3 pour novembre à mars)
                return current_month >= start or current_month <= end
        except (ValueError, TypeError):
            return False
    else:
        # Si pas de colonnes saison, utiliser les mois actifs détectés
        return current_month in SEASON_MONTHS

# Filtrer les paramètres pour ne garder que les variétés en saison
params_in_season = params[params.apply(is_variety_in_season, axis=1)].copy()

if params_in_season.empty:
    print(f"⚠️ Aucune variété n'est en saison de plantation pour le mois {current_month}.")
    print("   Aucune prévision ne sera générée.")
    # Créer un DataFrame vide avec les colonnes attendues
    df_pred = pd.DataFrame(columns=["date", "horizon", "parcelle", "variety", "temp_mean", "temp_min", "temp_max", "rain_mm", "humidity", "sun_hours", "shortwave_radiation"])
else:
    excluded_count = len(params) - len(params_in_season)
    if excluded_count > 0:
        excluded_varieties = params[~params.index.isin(params_in_season.index)]["variety"].tolist()
        print(f"🌱 {excluded_count} variété(s) exclue(s) (hors saison) : {', '.join(excluded_varieties)}")
    print(f"✅ {len(params_in_season)} variété(s) en saison pour le mois {current_month}.")

# === GÉNÉRATION DES COMBINAISONS PARCELLE / VARIÉTÉ ===
if params_in_season.empty:
    # Si aucune variété en saison, df_pred est déjà créé vide plus haut
    df_pred = pd.DataFrame(columns=["date", "horizon", "parcelle", "variety", "temp_mean", "temp_min", "temp_max", "rain_mm", "humidity", "sun_hours", "shortwave_radiation"])
else:
    rows = []
    for _, p in params_in_season.iterrows():
        for i, w in enumerate(df_forecast.itertuples(index=False), start=0):
            horizon = "J0" if i == 0 else f"J+{i}"
            rows.append({
                "date": w.date,
                "horizon": horizon,
                "temp_mean": w.temp_mean,
                "temp_min": w.temp_min,
                "temp_max": w.temp_max,
                "rain_mm": w.rain_mm,
                "humidity": w.humidity,
                "sun_hours": w.sun_hours,
                "shortwave_radiation": getattr(w, 'shortwave_radiation', np.nan),
                "parcelle": p["parcelle"],
                "variety": p["variety"]
            })
    df_pred = pd.DataFrame(rows)

# === AJOUT DE NB_PLANTS DEPUIS PLANTS_PAR_ANNEE ===
if df_pred.empty:
    print("ℹ️ Aucune prévision à générer (aucune variété en saison).")
else:
    print("🌱 Ajout des données de plants par année aux prévisions...")
try:
    if USE_DATA_LOADER:
        plants_par_annee = load_plants_par_annee()
    else:
        plants_par_annee = pd.read_excel(EXCEL_PATH, sheet_name="Plants_par_annee")
    if not plants_par_annee.empty and not df_pred.empty:
        # Nettoyage des données
        plants_par_annee["variety"] = plants_par_annee["variety"].astype(str).str.strip().str.lower()
        plants_par_annee["Année"] = plants_par_annee["Année"].astype(int)
        
        # Extraction de l'année depuis la date de prévision
        df_pred["year"] = df_pred["date"].dt.year
        
        # Fusion avec Plants_par_annee sur variety et année
        df_pred = df_pred.merge(
            plants_par_annee[["variety", "Année", "Nb_plants"]],
            left_on=["variety", "year"],
            right_on=["variety", "Année"],
            how="left"
        )
        
        # Renommer la colonne
        if "Nb_plants" in df_pred.columns:
            df_pred.rename(columns={"Nb_plants": "nb_plants"}, inplace=True)
            df_pred.drop(columns=["Année"], inplace=True, errors="ignore")
        
        # Gestion des valeurs manquantes (utiliser la valeur de l'année la plus récente disponible)
        if df_pred["nb_plants"].isna().any():
            missing_count = df_pred["nb_plants"].isna().sum()
            print(f"⚠️ Avertissement : {missing_count} prévisions n'ont pas de 'nb_plants' (année/variété non trouvée).")
            print("👉 Utilisation de la valeur de l'année la plus récente disponible par variété.")
            
            # Pour chaque variété, prendre la valeur de l'année la plus récente
            for variety in df_pred["variety"].unique():
                mask = (df_pred["variety"] == variety) & (df_pred["nb_plants"].isna())
                if mask.any():
                    # Chercher la valeur la plus récente pour cette variété
                    variety_data = plants_par_annee[plants_par_annee["variety"] == variety]
                    if not variety_data.empty:
                        latest_year = variety_data["Année"].max()
                        latest_nb_plants = variety_data[variety_data["Année"] == latest_year]["Nb_plants"].iloc[0]
                        df_pred.loc[mask, "nb_plants"] = latest_nb_plants
            
            # Si toujours NaN, utiliser la moyenne par variété
            default_by_variety = df_pred.groupby("variety")["nb_plants"].mean()
            for variety in default_by_variety.index:
                mask = (df_pred["variety"] == variety) & (df_pred["nb_plants"].isna())
                if not pd.isna(default_by_variety[variety]):
                    df_pred.loc[mask, "nb_plants"] = default_by_variety[variety]
            
            # Si toujours NaN, utiliser la moyenne globale
            df_pred["nb_plants"] = df_pred["nb_plants"].fillna(df_pred["nb_plants"].mean())
        
        print(f"✅ Données de plants ajoutées aux prévisions")
        # Supprimer la colonne temporaire "year" si elle existe
        df_pred.drop(columns=["year"], inplace=True, errors="ignore")
    else:
        if not df_pred.empty:
            print("⚠️ L'onglet 'Plants_par_annee' est vide. Colonne 'nb_plants' non ajoutée.")
            df_pred["nb_plants"] = np.nan
except Exception as e:
    if not df_pred.empty:
        print(f"⚠️ Erreur lors de la lecture de 'Plants_par_annee' : {e}")
        print("👉 Colonne 'nb_plants' non ajoutée.")
        df_pred["nb_plants"] = np.nan

# === AJOUT DES FEATURES D'ORGANISATION HEBDOMADAIRE ===
if df_pred.empty:
    print("ℹ️ Aucune feature hebdomadaire à ajouter (aucune prévision).")
else:
    print("📅 Ajout des features d'organisation hebdomadaire aux prévisions...")

    # Jour de la semaine (0=Lundi, 6=Dimanche)
    df_pred["jour_semaine"] = df_pred["date"].dt.dayofweek

    # Charger les paramètres de récolte quotidienne
    recolte_quotidienne = pd.DataFrame()  # Initialiser pour éviter les erreurs
    try:
        if USE_DATA_LOADER:
            recolte_quotidienne = load_recolte_quotidienne()
        else:
            recolte_quotidienne = pd.read_excel(EXCEL_PATH, sheet_name="Recolte_quotidienne")
        if not recolte_quotidienne.empty:
            # Créer un mapping jour_semaine_num -> fraction_fraiseraie
            fraction_map = dict(zip(
                recolte_quotidienne["jour_semaine_num"],
                recolte_quotidienne["fraction_fraiseraie"]
            ))
            # Appliquer le mapping
            df_pred["fraction_fraiseraie"] = df_pred["jour_semaine"].map(fraction_map)
            print(f"✅ Paramètres de récolte quotidienne chargés depuis Excel ({len(recolte_quotidienne)} jours)")
        else:
            recolte_quotidienne = pd.DataFrame()  # Réinitialiser si vide
            raise ValueError("L'onglet 'Recolte_quotidienne' est vide")
    except Exception as e:
        print(f"⚠️ Erreur lors de la lecture de 'Recolte_quotidienne' : {e}")
        print("👉 Utilisation des valeurs par défaut.")
        # Valeurs par défaut en cas d'erreur
        recolte_quotidienne = pd.DataFrame()  # S'assurer qu'elle est vide pour utiliser le fallback
        def get_fraction_fraiseraie(dayofweek):
            if dayofweek in [0, 1, 2]:  # Lundi, Mardi, Mercredi
                return 1/3
            elif dayofweek in [3, 4, 5]:  # Jeudi, Vendredi, Samedi
                return 1/2
            else:  # Dimanche
                return 0
        df_pred["fraction_fraiseraie"] = df_pred["jour_semaine"].apply(get_fraction_fraiseraie)

    # Calcul de jours_since_last_recolte pour chaque parcelle/variété
    # IMPORTANT : Prise en compte de l'inertie de la récolte basée sur les paramètres de récolte quotidienne
    # Pour chaque jour de récolte, la fraction récoltée n'a pas été récoltée depuis le dernier jour de récolte
    # dans le cycle hebdomadaire (selon les paramètres de récolte quotidienne)

    # Construire un mapping des jours de récolte depuis les paramètres
    if not recolte_quotidienne.empty:
        # Identifier les jours de récolte (fraction > 0)
        jours_recolte = recolte_quotidienne[recolte_quotidienne["fraction_fraiseraie"] > 0].copy()
        jours_recolte = jours_recolte.sort_values("jour_semaine_num")
        jours_recolte_nums = sorted(jours_recolte["jour_semaine_num"].tolist())
        
        # Créer un mapping : pour chaque jour, trouver le dernier jour de récolte avant lui dans le cycle
        def find_last_harvest_day_before(current_day):
            """Trouve le dernier jour de récolte avant current_day dans le cycle hebdomadaire."""
            # Chercher dans les jours précédents de la semaine actuelle
            prev_days = [d for d in jours_recolte_nums if d < current_day]
            if prev_days:
                return max(prev_days)
            # Si pas de jour précédent dans la semaine, prendre le dernier jour de récolte de la semaine précédente
            if jours_recolte_nums:
                return max(jours_recolte_nums)
            return None
        
        def calculate_jours_since_last_recolte_with_inertia(row, df_hist):
            """Calcule les jours depuis la dernière récolte en tenant compte de l'inertie hebdomadaire."""
            current_date = row["date"]
            current_dayofweek = current_date.dayofweek  # 0=Lundi, 6=Dimanche
            parcelle = row["parcelle"]
            variety = row["variety"]
            
            # Filtrer l'historique pour cette parcelle/variété
            hist_filtered = df_hist[(df_hist["parcelle"] == parcelle) & (df_hist["variety"] == variety)]
            
            if hist_filtered.empty:
                return 0
            
            # Si ce n'est pas un jour de récolte, utiliser la logique standard
            if current_dayofweek not in jours_recolte_nums:
                last_recolte_date = hist_filtered["date"].max()
                return (current_date - last_recolte_date).days
            
            # Trouver le dernier jour de récolte avant ce jour dans le cycle
            last_harvest_day = find_last_harvest_day_before(current_dayofweek)
            
            if last_harvest_day is None:
                return 0
            
            # Calculer les jours depuis ce dernier jour de récolte
            # Si le dernier jour est dans la semaine actuelle (avant aujourd'hui)
            if last_harvest_day < current_dayofweek:
                days_since = current_dayofweek - last_harvest_day
                # Trouver la date du dernier jour de récolte dans l'historique
                last_harvest_date = current_date - pd.Timedelta(days=days_since)
                prev_recoltes = hist_filtered[hist_filtered["date"] <= last_harvest_date]
                if not prev_recoltes.empty:
                    actual_last_date = prev_recoltes["date"].max()
                    return (current_date - actual_last_date).days
                return days_since
            else:
                # Le dernier jour de récolte est dans la semaine précédente
                # Calculer les jours depuis ce jour de la semaine précédente
                days_back_to_last_harvest = (7 - last_harvest_day) + current_dayofweek
                last_harvest_date = current_date - pd.Timedelta(days=days_back_to_last_harvest)
                prev_recoltes = hist_filtered[hist_filtered["date"] <= last_harvest_date]
                if not prev_recoltes.empty:
                    actual_last_date = prev_recoltes["date"].max()
                    return (current_date - actual_last_date).days
                return days_back_to_last_harvest
        
        if not df_hist.empty:
            df_pred["jours_since_last_recolte"] = df_pred.apply(
                lambda row: calculate_jours_since_last_recolte_with_inertia(row, df_hist),
                axis=1
            )
        else:
            df_pred["jours_since_last_recolte"] = 0
    else:
        # Fallback : utiliser la logique standard si pas de paramètres
        if not df_hist.empty:
            last_recolte_by_parcelle_variety = (
                df_hist.groupby(["parcelle", "variety"])["date"]
                .max()
                .reset_index()
                .rename(columns={"date": "last_recolte_date"})
            )
            df_pred = df_pred.merge(
                last_recolte_by_parcelle_variety,
                on=["parcelle", "variety"],
                how="left"
            )
            df_pred["jours_since_last_recolte"] = (
                (df_pred["date"] - df_pred["last_recolte_date"]).dt.days
            ).fillna(0)
            df_pred.drop(columns=["last_recolte_date"], inplace=True)
        else:
            df_pred["jours_since_last_recolte"] = 0

    # Calcul de jours_since_last_recolte_globale
    # On utilise la dernière date de récolte globale de l'historique
    if not df_hist.empty:
        last_global_date = df_hist["date"].max()
        df_pred["jours_since_last_recolte_globale"] = (
            (df_pred["date"] - last_global_date).dt.days
        )
    else:
        df_pred["jours_since_last_recolte_globale"] = 0

    print(f"✅ Features hebdomadaires ajoutées (jours: {sorted(df_pred['jour_semaine'].unique())})")

# === AJOUT DU CONTEXTE DE PRODUCTION ===
if not df_pred.empty:
    df_pred = df_pred.merge(valid_last[["parcelle", "variety", "kg_par_rangee_prev_day"]], on=["parcelle", "variety"], how="left")
    df_pred["kg_par_rangee_prev_day"] = df_pred["kg_par_rangee_prev_day"].fillna(0)

    # === Écrase par valeur du jour si disponible ===
    if not df_jour.empty and "kg_par_rangee" in df_jour.columns:
        df_pred = df_pred.merge(
            df_jour[["parcelle", "variety", "kg_par_rangee"]],
            on=["parcelle", "variety"],
            how="left",
            suffixes=("", "_today")
        )
        if "kg_par_rangee_today" in df_pred.columns:
            df_pred["kg_par_rangee_prev_day"] = df_pred["kg_par_rangee_today"].fillna(df_pred["kg_par_rangee_prev_day"])
            df_pred.drop(columns=["kg_par_rangee_today"], inplace=True)
    else:
        print("ℹ️ Aucun 'Jour_courant' renseigné — prévision basée uniquement sur les données historiques saisonnières.")

# === FEATURES DÉBUT/FIN DE SAISON (PRÉVISION) ===
if not df_pred.empty:
    # Première date de récolte de l'année courante
    df_hist_this_year = df_hist[df_hist["date"].dt.year == current_year]
    if not df_hist_this_year.empty:
        first_harvest_date = df_hist_this_year["date"].min()
    else:
        first_harvest_date = pd.Timestamp.now().normalize()

    df_pred["jours_depuis_premiere_recolte_annee"] = (df_pred["date"] - first_harvest_date).dt.days
    df_pred["jours_depuis_premiere_recolte_annee"] = df_pred["jours_depuis_premiere_recolte_annee"].clip(lower=0)

    # Moyenne 7j kg_par_rangee par (parcelle, variety) : derniers 7 points connus (historique + jour courant si présent)
    hist_kg = df_hist[["parcelle", "variety", "date", "kg_total", "nb_rangees"]].copy()
    hist_kg["kg_par_rangee"] = hist_kg["kg_total"] / hist_kg["nb_rangees"].replace(0, np.nan)
    hist_kg = hist_kg.dropna(subset=["kg_par_rangee"])

    today_norm = pd.Timestamp.now().normalize()
    mean_7j_list = []
    for _, row in df_pred[["parcelle", "variety"]].drop_duplicates().iterrows():
        parcelle, variety = row["parcelle"], row["variety"]
        sub = hist_kg[(hist_kg["parcelle"] == parcelle) & (hist_kg["variety"] == variety)].sort_values("date")
        vals = sub["kg_par_rangee"].tolist()
        if not df_jour.empty and "kg_par_rangee" in df_jour.columns:
            j = df_jour[(df_jour["parcelle"] == parcelle) & (df_jour["variety"] == variety)]
            if not j.empty and not pd.isna(j["kg_par_rangee"].iloc[0]):
                vals.append(j["kg_par_rangee"].iloc[0])
                vals = sorted(zip([*sub["date"].tolist(), today_norm], vals), key=lambda x: x[0])[-7:]
                vals = [v for _, v in vals]
            else:
                vals = vals[-7:]
        else:
            vals = vals[-7:]
        mean_7j = np.mean(vals) if vals else 0.0
        mean_7j_list.append({"parcelle": parcelle, "variety": variety, "moyenne_7j_kg_par_rangee": mean_7j})

    df_mean_7j = pd.DataFrame(mean_7j_list)
    df_pred = df_pred.merge(df_mean_7j, on=["parcelle", "variety"], how="left")
    df_pred["moyenne_7j_kg_par_rangee"] = df_pred["moyenne_7j_kg_par_rangee"].fillna(0)
    print("✅ Features 'jours_depuis_premiere_recolte_annee' et 'moyenne_7j_kg_par_rangee' ajoutées aux prévisions.")

# Sauvegarde temporaire des identifiants et données météo avant encodage
# Inclure les données météo dans l'export (elles sont déjà dans df_pred depuis la génération)
if not df_pred.empty:
    meteo_cols = ["temp_mean", "temp_min", "temp_max", "rain_mm", "humidity", "sun_hours", "shortwave_radiation"]
    cols_to_save = ["date", "horizon", "parcelle", "variety"]
    # Ajouter les colonnes météo si elles existent dans df_pred
    for col in meteo_cols:
        if col in df_pred.columns:
            cols_to_save.append(col)
    df_pred_for_output = df_pred[cols_to_save].copy()
else:
    df_pred_for_output = df_pred.copy()

# === ALIGNEMENT DES FEATURES ===
if df_pred.empty:
    print("ℹ️ Aucune prédiction à calculer (aucune variété en saison).")
    df_pred_for_output = df_pred.copy()
    predictions_kg_par_rangee = np.array([])
    predictions_std_par_rangee = np.array([])
    # Créer un DataFrame avec les colonnes attendues même si vide
    df_pred = pd.DataFrame(columns=[
        "date", "horizon", "parcelle", "variety",
        "temp_mean", "temp_min", "temp_max", "rain_mm", "humidity", "sun_hours", "shortwave_radiation",
        "predicted_kg_par_rangee", "predicted_std_par_rangee",
        "nb_rangees",
        "predicted_kg_total", "predicted_std_kg_total",
        "confidence_min_kg_total", "confidence_max_kg_total"
    ])
else:
    X_cols = model.feature_names_in_
    X_encoded = pd.get_dummies(df_pred, columns=["parcelle", "variety"], drop_first=True)
    for col in X_cols:
        if col not in X_encoded.columns:
            X_encoded[col] = 0

    X = X_encoded[X_cols]

    # === PRÉDICTION AVEC INTERVALLE DE CONFIANCE ===
    print("🧮 Calcul des prévisions avec intervalle de confiance...")
    all_tree_preds = np.stack([tree.predict(X) for tree in model.estimators_])
    predictions_kg_par_rangee = all_tree_preds.mean(axis=0)
    predictions_std_par_rangee = all_tree_preds.std(axis=0)

    # Ajouter les prédictions à df_pred_for_output
    df_pred_for_output["predicted_kg_par_rangee"] = predictions_kg_par_rangee
    df_pred_for_output["predicted_std_par_rangee"] = predictions_std_par_rangee

    # Merge pour récupérer nb_rangees et calcul du total
    # Récupérer aussi les données météo depuis df_pred (elles sont déjà dans df_pred_for_output)
    merged = df_pred_for_output.merge(params, on=["parcelle", "variety"], how="left")
    # Gérer le cas où nb_rangees est vide
    merged["nb_rangees"] = merged["nb_rangees"].fillna(10)  # Valeur par défaut

    # Ajouter les prédictions
    merged["predicted_kg_par_rangee"] = df_pred_for_output["predicted_kg_par_rangee"]
    merged["predicted_std_par_rangee"] = df_pred_for_output["predicted_std_par_rangee"]
    merged["predicted_kg_total"] = merged["predicted_kg_par_rangee"] * merged["nb_rangees"]
    merged["predicted_std_kg_total"] = merged["predicted_std_par_rangee"] * merged["nb_rangees"]
    merged["confidence_min_kg_total"] = (merged["predicted_kg_total"] - merged["predicted_std_kg_total"]).clip(lower=0)
    merged["confidence_max_kg_total"] = merged["predicted_kg_total"] + merged["predicted_std_kg_total"]

    # Lissage début de saison : réduire les prévisions pour les 7 premiers jours après la première récolte
    if "jours_depuis_premiere_recolte_annee" in df_pred.columns:
        merged["jours_depuis_premiere_recolte_annee"] = df_pred["jours_depuis_premiere_recolte_annee"].values
        ramp_days = 7
        scale = np.minimum(1.0, (merged["jours_depuis_premiere_recolte_annee"] + 1) / ramp_days)
        merged["predicted_kg_par_rangee"] = merged["predicted_kg_par_rangee"] * scale
        merged["predicted_std_par_rangee"] = merged["predicted_std_par_rangee"] * scale
        merged["predicted_kg_total"] = merged["predicted_kg_par_rangee"] * merged["nb_rangees"]
        merged["predicted_std_kg_total"] = merged["predicted_std_par_rangee"] * merged["nb_rangees"]
        merged["confidence_min_kg_total"] = (merged["predicted_kg_total"] - merged["predicted_std_kg_total"]).clip(lower=0)
        merged["confidence_max_kg_total"] = merged["predicted_kg_total"] + merged["predicted_std_kg_total"]
        merged.drop(columns=["jours_depuis_premiere_recolte_annee"], inplace=True, errors="ignore")
        print("✅ Lissage début de saison appliqué (rampe sur 7 jours).")

    df_pred = merged.copy()

# Réorganiser les colonnes pour un affichage plus clair
col_order = [
    "date", "horizon", "parcelle", "variety",
    "temp_mean", "temp_min", "temp_max", "rain_mm", "humidity", "sun_hours", "shortwave_radiation",
    "predicted_kg_par_rangee", "predicted_std_par_rangee",
    "nb_rangees",
    "predicted_kg_total", "predicted_std_kg_total",
    "confidence_min_kg_total", "confidence_max_kg_total"
]
# Garder seulement les colonnes qui existent
col_order = [col for col in col_order if col in df_pred.columns]
# Ajouter les colonnes restantes
remaining_cols = [col for col in df_pred.columns if col not in col_order]
df_pred = df_pred[col_order + remaining_cols]

if not df_pred.empty:
    df_pred = df_pred.sort_values(["date", "horizon", "parcelle", "variety"]).reset_index(drop=True)

# === SAUVEGARDE DANS LA BASE DE DONNÉES ===
try:
    from database import save_forecast, export_forecast_to_excel
    USE_DB = True
except ImportError:
    USE_DB = False
    print("⚠️ Module database non disponible, export Excel uniquement")

forecast_date = datetime.now().strftime("%Y-%m-%d")

if USE_DB:
    if df_pred.empty:
        print("\nℹ️ Aucune prévision à enregistrer (aucune variété en saison). Base de données non mise à jour.")
    else:
        # Préparer le DataFrame pour la base de données
        # S'assurer que la colonne 'date' est au format string pour la DB
        df_pred_db = df_pred.copy()
        if 'date' in df_pred_db.columns:
            if pd.api.types.is_datetime64_any_dtype(df_pred_db['date']):
                df_pred_db['date'] = df_pred_db['date'].dt.strftime('%Y-%m-%d')
            else:
                df_pred_db['date'] = df_pred_db['date'].astype(str)
        
        # Mapper les colonnes météo vers les noms attendus par la base de données
        column_mapping = {
            'temp_max': 'temperature_max',
            'temp_min': 'temperature_min',
            'rain_mm': 'precipitation_sum',
            'sun_hours': 'sunshine_duration',
            'humidity': 'relative_humidity_mean',
            'shortwave_radiation': 'shortwave_radiation_sum'
        }
        
        # Renommer les colonnes si elles existent
        for old_name, new_name in column_mapping.items():
            if old_name in df_pred_db.columns:
                df_pred_db[new_name] = df_pred_db[old_name]
        
        # Convertir sun_hours de heures en secondes si nécessaire (la DB stocke en secondes)
        if 'sunshine_duration' in df_pred_db.columns:
            # Si la valeur est < 24, c'est probablement en heures, convertir en secondes
            df_pred_db['sunshine_duration'] = df_pred_db['sunshine_duration'].apply(
                lambda x: x * 3600.0 if x and x < 24 else x
            )
        
        # Sauvegarder dans la base de données
        count = save_forecast(forecast_date, df_pred_db)
        print(f"\n💾 {count} prévisions sauvegardées dans la base de données (date: {forecast_date})")
        
        # Optionnel : exporter aussi en Excel pour compatibilité
        try:
            output_path = export_forecast_to_excel(forecast_date=forecast_date)
            print(f"📦 Prévisions également exportées en Excel : {output_path}")
        except Exception as e:
            print(f"⚠️ Export Excel optionnel échoué : {e}")
else:
    # Fallback : export Excel uniquement (ancien comportement)
    if df_pred.empty:
        print("\nℹ️ Aucune prévision à exporter (aucune variété en saison).")
    else:
        FORECASTS_DIR.mkdir(exist_ok=True)
        output_path = FORECASTS_DIR / f"forecast_week_{forecast_date}.xlsx"
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_pred.to_excel(writer, index=False, sheet_name="Prévisions")
        
        # Mise en forme auto (colonnes élargies)
        wb = load_workbook(output_path)
        ws = wb["Prévisions"]
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
        
        # Style d'en-tête (gras + fond gris clair)
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        
        wb.save(output_path)
        print(f"\n📦 Prévisions exportées : {output_path}")

# === RÉSUMÉ JOURNALIER EN CONSOLE ===
if df_pred.empty:
    print("\n📅 RÉSUMÉ DES PRÉVISIONS :")
    print("   Aucune prévision générée (aucune variété en saison de plantation).")
else:
    summary = (
        df_pred.groupby(["date", "horizon"])
        .agg(
            total_kg=("predicted_kg_total", "sum"),
            min_conf=("confidence_min_kg_total", "sum"),
            max_conf=("confidence_max_kg_total", "sum")
        )
        .reset_index()
    )

    print("\n📅 RÉSUMÉ DES PRÉVISIONS :")
    for _, row in summary.iterrows():
        print(
            f" - {row['date'].date()} ({row['horizon']}) → "
            f"{row['total_kg']:.1f} kg  "
            f"(min {row['min_conf']:.1f} / max {row['max_conf']:.1f})"
        )

print("✅ Terminé avec succès.")
